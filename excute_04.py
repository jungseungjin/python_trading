

import requests
import authorization_token
import os
import openpyxl
import datetime
import json
import time as timer

#엑셀에 저장되어있는 값을 기준으로 백테스팅을 시작한다
#1번 알고리즘
#볼린저밴드를 이용.
#볼린저밴드의 상단보다 지금의 종가가 높으면 상승신호로 판단을 하고 매수를 시작한다.
execlFileNames = os.listdir('./excel')
for excelFileName in execlFileNames:
    excelName = f'./excel/{excelFileName}'
    if(excelName == "./excel/.DS_Store"):
        continue
    excel= openpyxl.load_workbook(excelName)
    sheetNames = excel.sheetnames
    print(excelName)
    for sheetName in sheetNames:
        if sheetName == "Sheet" : continue
        market = sheetName
        sheet = excel[f"{market}"]
        maxRow = sheet.max_row
        money = 100000 #자본
        coin = 0 #코인갯수
        averagePrice = 0 #매수 평균가
        fee = 0.05#수수료는 거래대금의 퍼센티지 매 거래마다 곱해서 제한다
        totalFee = 0#전체 수수료
        totalSum = 0#전체 매수금액
        totalSell = 0#전체 매도금액
        buyPercent = 1.5#익절은 매수가 대비 buyPercent이상일때
        sellPercent = 0.5#손절은 매수가 대비 sellPercent이하일때
        #P 라인에 현재 자본을 입력한다
        #코인의 갯수는 고려하지 않음
        #한번에 매수하는 값은 1000
        #모두 매수를 했을 경우 기회가 오기 전까지 매도하지 않음
        for  i in range(26,maxRow):
            #거래가 발생되면 다음 분봉으로 넘어간다
            #매수 시나리오
            lastPrice = float(sheet[f"P{i}"].value)
            startPrice = float(sheet[f"B{i}"].value)
            bTop = float(sheet[f"K{i}"].value)
            nextStartPrice = float(sheet[f"B{i+1}"].value)
            if lastPrice > bTop :
                if money > 1000 : 
                    coin = coin+float(1000/nextStartPrice)
                    money = money - (1000+1000*0.05)
                    totalFee = totalFee + (1000*0.05)
                    totalBuy = totalBuy + 1000
                    averagePrice = float(1000/float(1000/nextStartPrice))
                    sheet[f"P{i}"] = money
                    continue
                
            #매도 시나리오
            #코인이 있을때 
            if coin > 0 : #매수평균가가 시가 혹은 종가보다 1.5%이상 높으면 매도(익절)
                if averagePrice*1.015 > startPrice or averagePrice*1.015 > lastPrice: 
                    targetPrice = 0
                    if averagePrice*1.015 > startPrice : 
                        targetPrice = startPrice
                    if averagePrice*1.015 > lastPrice : 
                        targetPrice = lastPrice
                    money = money + (coin*targetPrice - coin*targetPrice*0.05)
                    totalFee = totalFee + (coin*targetPrice*0.05)
                    totalSell = totalSell + (coin*targetPrice - coin*targetPrice*0.05)
                    coin = 0
                    averagePrice = 0
                elif averagePrice*0.995 < lastPrice : #매수평균가가 시가 혹은 종가보다 0.5%이상 낮으면 매도 (손절)


            






#B1 시가 C1고가 D1저가 E1종가
#H1 단기 이동평균선 5개  
#I1 중기 이동평균선 10개
#J1 장기 이동평균선 20개
#K1 볼린저밴드 상단 - 10개 이동평균선 + (10개 표준편차 * 곱)
#L1 볼린저밴드 중간 - 이동평균선 10개
#M1 볼린저밴드 하단 - 10개 이동평균선 - (10개 표준편차 * 곱)
#N1 종가-시가(캔들의 길이)
#O1 고가-저가(꼬리의 길이)
