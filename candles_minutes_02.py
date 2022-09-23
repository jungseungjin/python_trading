# Response
# 필드	설명	타입
# market	마켓명	String
# candle_date_time_utc	캔들 기준 시각(UTC 기준)
# 포맷: yyyy-MM-dd'T'HH:mm:ss	String
# candle_date_time_kst	캔들 기준 시각(KST 기준)
# 포맷: yyyy-MM-dd'T'HH:mm:ss	String
# opening_price	시가	Double
# high_price	고가	Double
# low_price	저가	Double
# trade_price	종가	Double
# timestamp	해당 캔들에서 마지막 틱이 저장된 시각	Long
# candle_acc_trade_price	누적 거래 금액	Double
# candle_acc_trade_volume	누적 거래량	Double
# unit	분 단위(유닛)	Integer

# PATH PARAMS
# unit
# int32
# required
# 분 단위. 가능한 값 : 1, 3, 5, 15, 10, 30, 60, 240

# 1
# QUERY PARAMS
# market
# string
# required
# 마켓 코드 (ex. KRW-BTC)

# KRW-BTC
# to
# string
# 마지막 캔들 시각 (exclusive). 포맷 : yyyy-MM-dd'T'HH:mm:ss'Z' or yyyy-MM-dd HH:mm:ss. 비워서 요청시 가장 최근 캔들

# count
# int32
# 캔들 개수(최대 200개까지 요청 가능)

# 1


import requests
import authorization_token
import os
import openpyxl
import datetime
import json
import time as timer

unit = 5
RequestCount = 200
headers = {"accept": "application/json"}

execlFileNames = os.listdir('./excel')
for excelFileName in execlFileNames:
    excelName = f'./excel/{excelFileName}'
    if(excelName == "./excel/.DS_Store"):
        continue
    if(excelName != "./excel/KRW-GMT.xlsx"):
        continue
    excel= openpyxl.load_workbook(excelName)
    sheetNames = excel.sheetnames
    print(excelName)
    for sheetName in sheetNames:
        if sheetName == "Sheet" : continue
        market = sheetName
        sheet = excel[f"{market}"]
        maxRow = sheet.max_row
        #1열2행값 가져오기 -> 값이 비어있으면 새로 시작하는것 -> time = datetime.datetime.now()
        #값이 비어있지 않으면 -> 그것이 time값
        if maxRow == 1 : 
            sheet["A1"] = "시간"
            sheet["B1"] = "시가"
            sheet["C1"] = "고가"
            sheet["D1"] = "저가"
            sheet["E1"] = "종가"
            sheet["F1"] = "누적 거래금액"
            sheet["G1"] = "누적 거래량"
            sheet["H1"] = "단기 이동평균선"
            sheet["I1"] = "중기 이동평균선"
            sheet["J1"] = "장기 이동평균선"
            sheet["K1"] = "볼린저밴드 상단"
            sheet["L1"] = "볼린저밴드 중간"
            sheet["M1"] = "볼린저밴드 하단"
            sheet["N1"] = "종가-시가(캔들의 길이)"
            sheet["O1"] = "고가-저가(꼬리의 길이)"
            sheet["P1"] = "방법1"
            sheet["Q1"] = "방법2"
            sheet["R1"] = "방법3"
            sheet["S1"] = "방법4"
            sheet["T1"] = "방법5"
            
            time = datetime.datetime.now()
            time = str(time)[:-7]
            time = datetime.datetime.strptime(time,"%Y-%m-%d %H:%M:%S")
            # excel.save(excelName)
        else :
            A2Value = str(sheet[f"A{maxRow}"].value)
            time = datetime.datetime.strptime(A2Value,"%Y-%m-%dT%H:%M:%S")
            time = time - datetime.timedelta(hours=9)
        while True : 
            timeFormat = str(time)
            url = f"https://api.upbit.com/v1/candles/minutes/{unit}?market={market}&count={RequestCount}&to={timeFormat}"
            response = requests.get(url, headers=headers)
            resultData = json.loads(response.text)
            
            if len(resultData) == 0 : 
                break
            resultData.reverse()
            sheet.insert_rows(2,len(resultData))
            time = resultData[0]['candle_date_time_utc']
            for index,row in enumerate(resultData) :
                sheet[f"A{index+2}"] = row['candle_date_time_kst']
                sheet[f"B{index+2}"] = row['opening_price']
                sheet[f"C{index+2}"] = row['high_price']
                sheet[f"D{index+2}"] = row['low_price']
                sheet[f"E{index+2}"] = row['trade_price']
                sheet[f"F{index+2}"] = row['candle_acc_trade_price']
                sheet[f"G{index+2}"] = row['candle_acc_trade_volume']
            print(sheet.max_row)
            timer.sleep(0.1)

        excel.save(excelName)


# for sheetName in sheetNames:
#     market = sheetName



#배열의 마지막에 나오는 candle_date_time_kst 를 형식변환해서 to에 넣어준다.
"""
엑셀시트에 넣을때는
1. 해당하는 엑셀 시트에서 0번째 얄값을 모두 가져온다.
2. 0번째 열에서 가장 값이 낮은 행값을 가져온다.(시간기준) candle_date_time_kst
3. 옛날데이터가 위로 최신데이터는 아래로 간다.
4. 옛날날짜 기준으로 리퀘스트 실행해서 값을 가져온다
5. 가져온 값의 순서를 반대로 돌리고 엑셀시트 모양에 맞게 형태를 맞춘다
6. 엑셀의 1행부터 순서대로 대입한다.
7. candle_date_time_kst가 겹칠경우 어떻게 할것인가?
"""