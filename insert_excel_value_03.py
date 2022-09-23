

from ast import Num
import requests
import authorization_token
import os
import openpyxl
import datetime
import json
import time as timer
import numpy

#주어진 값으로 새로운 값을 계산해서 새로운 셀에 넣는다.
#넣을 값 : 이동평균선 , 볼린저밴드 , 이전 데이터의 종가-시가(캔들의 길이) , 이전 데이터의 고가-저가(꼬리의 길이), 매수량, 매도량
#B1 시가 C1고가 D1저가 E1종가 F1누적 거래금액 G1누적거래량  지금은 거래금액 거래량 바뀌었음;;  
#볼린저밴드는 보통 20일치를 평균으로 보지만 여기서는 10으로 기준으로 해본다
#numpy.mean(list) - 평균
#numpy.std(list) - 표준편차

#H1 단기 이동평균선 5개  
#I1 중기 이동평균선 10개
#J1 장기 이동평균선 20개
#K1 볼린저밴드 상단 - 10개 이동평균선 + (10개 표준편차 * 곱)
#L1 볼린저밴드 중간 - 이동평균선 10개
#M1 볼린저밴드 하단 - 10개 이동평균선 - (10개 표준편차 * 곱)
#N1 종가-시가(캔들의 길이)
#O1 고가-저가(꼬리의 길이)
#P1 Q1 R1 S1 T1 U1 V1 W1 X1 Y1 Z1

execlFileNames = os.listdir('./excel')
for excelFileName in execlFileNames:
    excelName = f'./excel/{excelFileName}'
    if(excelName == "./excel/.DS_Store"):
        continue
    excel= openpyxl.load_workbook(excelName)
    sheetNames = excel.sheetnames
    print(excelName)
    for sheetName in sheetNames:
        if sheetName == "Sheet" : continue
        market = sheetName
        sheet = excel[f"{market}"]
        maxRow = sheet.max_row
            
        sum5 = 0
        sum10 = 0
        stdList = []
        sum20 = 0

        for rowNumber in range(25,maxRow) : 
            if(rowNumber == 25):
                for index in range(21,25) : 
                    sum5 = sum5 + float(sheet[f"E{index}"].value)
                for index in range(16,25) : 
                    sum10 = sum10 + float(sheet[f"E{index}"].value)
                    stdList.append(float(sheet[f"E{index}"].value))
                for index in range(6,25) : 
                    sum20 = sum20 + float(sheet[f"E{index}"].value)
            # 여기서부터 값 계산 시작
            
            std10 = numpy.std(stdList)
            H1 = sum5/5
            I1 = sum10/10
            J1 = sum20/20
            K1 = I1 + (std10*2)
            L1 = I1
            M1 = I1 - (std10*2)
            N1 = float(sheet[f"E{rowNumber}"].value) - float(sheet[f"B{rowNumber}"].value)
            O1 = float(sheet[f"C{rowNumber}"].value) - float(sheet[f"D{rowNumber}"].value)
            sheet[f"H{rowNumber}"] = H1
            sheet[f"I{rowNumber}"] = I1
            sheet[f"J{rowNumber}"] = J1
            sheet[f"K{rowNumber}"] = K1
            sheet[f"L{rowNumber}"] = L1
            sheet[f"M{rowNumber}"] = M1
            sheet[f"N{rowNumber}"] = N1
            sheet[f"O{rowNumber}"] = O1

            sum5 = sum5 - float(sheet[f"E{rowNumber - 4}"].value) + float(sheet[f"E{rowNumber + 1}"].value)
            sum10 = sum10 - float(sheet[f"E{rowNumber - 9}"].value) + float(sheet[f"E{rowNumber + 1}"].value)
            sum20 = sum20 - float(sheet[f"E{rowNumber - 19}"].value) + float(sheet[f"E{rowNumber + 1}"].value)
            del stdList[0]
            stdList.append(float(sheet[f"E{rowNumber + 1}"].value))
            if rowNumber%1000 == 0:
                print(rowNumber)
        
        excel.save(excelName)   

#B1 시가 C1고가 D1저가 E1종가
#H1 단기 이동평균선 5개  
#I1 중기 이동평균선 10개
#J1 장기 이동평균선 20개
#K1 볼린저밴드 상단 - 10개 이동평균선 + (10개 표준편차 * 곱)
#L1 볼린저밴드 중간 - 이동평균선 10개
#M1 볼린저밴드 하단 - 10개 이동평균선 - (10개 표준편차 * 곱)
#N1 종가-시가(캔들의 길이)
#O1 고가-저가(꼬리의 길이)




        